"""
DAY 6 — Replenishment Planner
==============================
Answers the question SAP APO costs $200K+ to answer:
  "When should I order, how much, and what does it cost?"

Inputs (from existing pipeline):
  - planning_database/inventory_risk.csv   (current stock + forecast demand)
  - planning_database/forecasts.csv        (monthly forecast per SKU x region)

New inputs (simulated — replace with real data when available):
  - Supplier lead times per category
  - Minimum Order Quantities (MOQs) per SKU
  - Safety stock policy (weeks of cover)

Outputs:
  - planning_database/replenishment_plan.csv     (machine-readable)
  - planning_database/replenishment_plan.xlsx    (planner-friendly, colour-coded)

Logic:
  1. Project inventory forward month by month using the forecast
  2. Detect the first month inventory drops below safety stock threshold
  3. Count back by lead time to get the ORDER DATE (today if already overdue)
  4. Calculate order quantity: bring stock back to target cover level
  5. Round up to nearest MOQ
  6. Flag urgency: Overdue / This Week / This Month / Planned
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_STYLING = True
except ImportError:
    EXCEL_STYLING = False

# ── Configuration (replace with real values when available) ───────────────
TODAY         = datetime(2024, 12, 1)   # Matches end of our historical data
TARGET_COVER  = 3.0                      # Target months of stock after replenishment
SAFETY_STOCK  = 1.5                      # Trigger reorder when coverage drops below this

# Supplier lead times (weeks) by category — realistic for outdoor gear
LEAD_TIMES = {
    'Sleeping Bags': 16,   # Made in Asia, complex manufacturing
    'Packs':         14,   # Frames + fabric, longer production
    'Towels':         8,   # Simpler textile, shorter lead
    'Accessories':   10,   # Mixed complexity
    'Camp':          12,   # Metal + fabric components
}

# Minimum Order Quantities by SKU — realistic for a mid-size outdoor brand
MOQS = {
    'STS-SL-001': 50,  'STS-SL-002': 50,  'STS-SL-003': 30,
    'STS-PK-001': 24,  'STS-PK-002': 48,  'STS-PK-003': 36,  'STS-PK-004': 12,
    'STS-TW-001': 100, 'STS-TW-002': 100, 'STS-TW-003': 72,
    'STS-AC-001': 200, 'STS-AC-002': 150, 'STS-AC-003': 150,
    'STS-CP-001': 24,  'STS-CP-002': 60,
}

# Cost of goods (approximate % of retail price) for order value calculation
COGS_PCT = 0.40

print("=" * 65)
print("  REPLENISHMENT PLANNER — Open-Source S&OP Engine")
print(f"  Run date : {TODAY.strftime('%d %b %Y')}")
print(f"  Target cover    : {TARGET_COVER} months after order arrival")
print(f"  Safety stock    : {SAFETY_STOCK} months (reorder trigger)")
print("=" * 65)

# ── Load data ─────────────────────────────────────────────────────────────
inv    = pd.read_csv('planning_database/inventory_risk.csv')
fcast  = pd.read_csv('planning_database/forecasts.csv')
fcast['date'] = pd.to_datetime(fcast['date'])

print(f"\n  Loaded {len(inv)} SKU-region positions")
print(f"  Loaded {len(fcast)} forecast data points")

# ── Core replenishment logic ───────────────────────────────────────────────
results = []

for _, row in inv.iterrows():
    sku      = row['sku']
    region   = row['region']
    pname    = row['product_name']
    category = row['category']
    price    = row['price_usd']

    # Get supplier lead time and MOQ
    lead_weeks = LEAD_TIMES.get(category, 12)
    lead_days  = lead_weeks * 7
    moq        = MOQS.get(sku, 50)

    # Get monthly forecasts for this SKU-region
    sku_fcast = (fcast[(fcast['sku']==sku) & (fcast['region']==region)]
                 .sort_values('date')[['date','forecast_units']])

    if sku_fcast.empty:
        continue

    # ── Step 1: Project inventory forward month by month ──────────────
    stock      = float(row['current_inventory'])
    stockout_month = None
    below_safety   = None
    projected_rows = []

    for _, fm in sku_fcast.iterrows():
        demand = float(fm['forecast_units'])
        stock_before = stock
        stock = max(0, stock - demand)
        monthly_avg  = float(row['forecast_monthly_avg'])
        cover_after  = stock / monthly_avg if monthly_avg > 0 else 99

        projected_rows.append({
            'month':         fm['date'],
            'demand':        demand,
            'stock_end':     round(stock, 0),
            'coverage_months': round(cover_after, 1),
        })

        if stock <= 0 and stockout_month is None:
            stockout_month = fm['date']
        if cover_after < SAFETY_STOCK and below_safety is None:
            below_safety = fm['date']

    # ── Step 2: Determine if replenishment is needed ───────────────────
    current_cover = float(row['coverage_months']) if not pd.isna(row['coverage_months']) else 0
    needs_order   = current_cover < TARGET_COVER or below_safety is not None

    if not needs_order:
        # Still calculate a planned future order date
        trigger_date  = TODAY + timedelta(days=int((current_cover - SAFETY_STOCK) * 30))
        order_date    = trigger_date - timedelta(days=lead_days)
        urgency       = 'Planned'
        order_needed  = False
    else:
        # Trigger date = when stock first drops below safety stock
        if below_safety:
            trigger_date = below_safety
        else:
            trigger_date = TODAY  # Already below safety stock

        order_date = trigger_date - timedelta(days=lead_days)
        order_needed = True

        # Determine urgency
        days_to_order = (order_date - TODAY).days
        if days_to_order < 0:
            urgency = 'OVERDUE'
        elif days_to_order <= 7:
            urgency = 'This Week'
        elif days_to_order <= 30:
            urgency = 'This Month'
        else:
            urgency = 'Upcoming'

    # ── Step 3: Calculate order quantity ──────────────────────────────
    monthly_demand = float(row['forecast_monthly_avg'])

    # Target ending stock = TARGET_COVER months of demand
    target_stock   = monthly_demand * TARGET_COVER
    # Net requirement = target - what we'll have when order arrives
    # Approximate stock at arrival = current - (lead_weeks/4) months of demand
    months_until_arrival = lead_weeks / 4
    projected_stock_at_arrival = max(0, float(row['current_inventory']) - monthly_demand * months_until_arrival)
    raw_order_qty = max(0, target_stock - projected_stock_at_arrival)

    # Round up to nearest MOQ
    if raw_order_qty > 0:
        order_qty = int(np.ceil(raw_order_qty / moq) * moq)
    else:
        order_qty = moq if needs_order else 0

    # Order value
    order_value_usd = round(order_qty * price * COGS_PCT, 2)

    # Arrival date
    arrival_date = order_date + timedelta(days=lead_days)

    results.append({
        'sku':                  sku,
        'product_name':         pname,
        'region':               region,
        'category':             category,
        'current_stock':        int(row['current_inventory']),
        'monthly_demand':       round(monthly_demand, 1),
        'coverage_months':      round(current_cover, 1),
        'risk_level':           row['risk_level'],
        'lead_time_weeks':      lead_weeks,
        'moq':                  moq,
        'safety_stock_months':  SAFETY_STOCK,
        'target_cover_months':  TARGET_COVER,
        'order_needed':         order_needed,
        'urgency':              urgency,
        'order_date':           order_date.strftime('%Y-%m-%d'),
        'arrival_date':         arrival_date.strftime('%Y-%m-%d'),
        'order_qty':            order_qty,
        'order_value_usd':      order_value_usd,
        'projected_stockout':   stockout_month.strftime('%Y-%m-%d') if stockout_month else 'No stockout',
        'below_safety_date':    below_safety.strftime('%Y-%m-%d') if below_safety else 'Never',
    })

plan = pd.DataFrame(results)

# ── Sort by urgency then coverage ─────────────────────────────────────────
urgency_order = {'OVERDUE':0,'This Week':1,'This Month':2,'Upcoming':3,'Planned':4}
plan['urgency_rank'] = plan['urgency'].map(urgency_order)
plan = plan.sort_values(['urgency_rank','coverage_months']).drop('urgency_rank', axis=1)

# ── Save CSV ───────────────────────────────────────────────────────────────
plan.to_csv('planning_database/replenishment_plan.csv', index=False)

# ── Save Excel with colour-coding ─────────────────────────────────────────
export_cols = [
    'sku','product_name','region','category',
    'current_stock','monthly_demand','coverage_months','risk_level',
    'lead_time_weeks','moq',
    'urgency','order_date','arrival_date',
    'order_qty','order_value_usd',
    'projected_stockout','below_safety_date',
]

plan[export_cols].to_excel(
    'planning_database/replenishment_plan.xlsx',
    index=False, sheet_name='Replenishment Plan'
)

if EXCEL_STYLING:
    wb = load_workbook('planning_database/replenishment_plan.xlsx')
    ws = wb.active

    # Colour fills
    fills = {
        'OVERDUE':    PatternFill('solid', fgColor='F4CCCC'),
        'This Week':  PatternFill('solid', fgColor='FCE5CD'),
        'This Month': PatternFill('solid', fgColor='FFF2CC'),
        'Upcoming':   PatternFill('solid', fgColor='D9EAD3'),
        'Planned':    PatternFill('solid', fgColor='EAF4FB'),
    }
    header_fill = PatternFill('solid', fgColor='1A4F8A')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Style header row
    for cell in ws[1]:
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 30

    # Find urgency column index
    headers = [cell.value for cell in ws[1]]
    urgency_col = headers.index('urgency') + 1

    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        urgency_val = row[urgency_col - 1].value
        fill = fills.get(urgency_val, PatternFill())
        for cell in row:
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    # Auto-width columns
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save('planning_database/replenishment_plan.xlsx')
    print("\n  Excel file colour-coded and formatted")

# ── Print summary ──────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print("  REPLENISHMENT SUMMARY")
print(f"{'='*65}")

by_urgency = plan.groupby('urgency').agg(
    skus=('sku','count'),
    total_units=('order_qty','sum'),
    total_value=('order_value_usd','sum')
).reindex(['OVERDUE','This Week','This Month','Upcoming','Planned'])

for urgency, row2 in by_urgency.iterrows():
    if pd.isna(row2['skus']): continue
    print(f"\n  {urgency}")
    print(f"    SKU-regions : {int(row2['skus'])}")
    print(f"    Units to order: {int(row2['total_units']):,}")
    print(f"    Order value : ${row2['total_value']:,.0f} USD (COGS)")

print(f"\n{'─'*65}")
total_val = plan[plan['order_needed']==True]['order_value_usd'].sum()
total_qty = plan[plan['order_needed']==True]['order_qty'].sum()
print(f"  TOTAL orders needed   : {plan['order_needed'].sum()} SKU-regions")
print(f"  TOTAL units to order  : {int(total_qty):,}")
print(f"  TOTAL order value     : ${total_val:,.0f} USD (at {int(COGS_PCT*100)}% COGS)")

print(f"\n  Files saved:")
print(f"    planning_database/replenishment_plan.csv")
print(f"    planning_database/replenishment_plan.xlsx")
print(f"\n  → Run day7_sop_pack.py next to generate the S&OP meeting pack.")
print(f"{'='*65}")
